import os
import django
import openpyxl

from core.models import Offer

# Si utilisé en script indépendant, décommente ci-dessous
# os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ton_projet.settings')
# django.setup()

##### Section update ###############
# Liste des sections valides (les "values" de SECTION_CHOICES)
VALID_SECTIONS = {
    'artists_groups',
    'courses_training',
    'calls_events',
    'services_equipment',
}

# Remplace par le chemin réel vers ton fichier
excel_path = 'pat_section.xlsx'

# Charger le fichier Excel
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

rows = list(ws.iter_rows(min_row=2, values_only=True))

updated = 0
not_found = 0
invalid_sections = []

for offer_id, section in rows:
    if section not in VALID_SECTIONS:
        invalid_sections.append((offer_id, section))
        continue

    try:
        offer = Offer.objects.get(id=offer_id)
        offer.section = section
        offer.save()
        updated += 1
    except Offer.DoesNotExist:
        print(f"❌ Offer ID {offer_id} introuvable.")
        not_found += 1

print(f"✅ {updated} offres mises à jour.")
if not_found:
    print(f"⚠️ {not_found} offres non trouvées.")
if invalid_sections:
    print(f"❗ {len(invalid_sections)} lignes ignorées à cause d'une section invalide :")
    for oid, sec in invalid_sections:
        print(f"  - ID {oid} → section invalide: '{sec}'")


######## Type & Category update ####################
# Valeurs valides (doivent correspondre aux "values" des CHOICES dans le modèle)
VALID_CATEGORIES = {'paid', 'unpaid'}
VALID_TYPES = {'offer', 'demand'}

# Chemin du fichier Excel
excel_path = 'pat_type_remuneration.xlsx'

# Charger le fichier
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Lire les lignes (on suppose que la 1ère est l’en-tête)
rows = list(ws.iter_rows(min_row=2, values_only=True))

updated = 0
not_found = 0
invalid_entries = []

for offer_id, category, offer_type in rows:
    # Validation des valeurs
    if category not in VALID_CATEGORIES or offer_type not in VALID_TYPES:
        invalid_entries.append((offer_id, category, offer_type))
        continue

    try:
        offer = Offer.objects.get(id=offer_id)
        offer.category = category
        offer.type = offer_type
        offer.save()
        updated += 1
    except Offer.DoesNotExist:
        print(f"❌ Offer ID {offer_id} introuvable.")
        not_found += 1

# Résumé
print(f"✅ {updated} offres mises à jour.")
if not_found:
    print(f"⚠️ {not_found} offres non trouvées.")
if invalid_entries:
    print(f"❗ {len(invalid_entries)} lignes ignorées à cause de valeurs invalides :")
    for oid, cat, typ in invalid_entries:
        print(f"  - ID {oid} → category: '{cat}', offer_type: '{typ}'")